#Andrew Bruckbauer
#12/20/2022
#Meant to be used to add a letter value to a colored cell in an excel spreadsheet for use in other scripts.

import openpyxl
from tqdm import tqdm  # tqdm is a library for creating progress bars
import os  # os is the Python standard library for interacting with the operating system

# Get the filename from the input
filename = input('Enter the filename: ')

# Open the Excel workbook
workbook = openpyxl.load_workbook(filename)

# Initialize the total number of cells to process
total_cells = 0

# Iterate through each sheet in the workbook
for sheet in workbook:
    # Add the number of cells in this sheet to the total
    total_cells += sheet.max_row * sheet.max_column

# Initialize a progress bar
pbar = tqdm(total=total_cells)

# Iterate through each sheet in the workbook
for sheet in workbook:
    # Iterate through each cell in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell's value is a letter
            #if isinstance(cell.value, str) and cell.value.isalpha():
            if cell.value and (cell.font.bold or cell.font.italic or cell.font.underline):
                # Skip this cell if it contains a letter
                continue
            # Check the cell's fill color
            if cell.fill.fgColor.rgb in ['FFFF0000', 'FFA61C00', 'FF990000', 'FF85200C', 'FF980000']:  # Red
                cell.value = 'R'
            elif cell.fill.fgColor.rgb in ['FF073763', 'FF351C75', 'FF1C4587']:  # Blue
                cell.value = 'B'
            elif cell.fill.fgColor.rgb == 'FF38761D':  # Green
                cell.value = 'G'
            elif cell.fill.fgColor.rgb in ['FFFFFF00', 'FFBF9000', 'FFF1C232']:  # Yellow
                cell.value = 'Y'
            # Update the progress bar
            pbar.update(1)

# Create the new filename
new_filename = filename.rsplit('.', 1)[0] + '_formatted.xlsx'

# Check if the new file already exists
i = 1
while os.path.exists(new_filename):
    # If the file exists, add a number to the end of the filename
    new_filename = f'{filename.rsplit(".", 1)[0]}_formatted_{i}.xlsx'
    i += 1

# Save the changes to a new workbook
workbook.save(new_filename)

# Close the progress bar
pbar.close()
